"""Validation for the Report Center — the library, and the workflow around it.

FIVE THINGS ARE BEING DEFENDED.

GENERATE IS NOT DOWNLOAD. `POST /api/reports` answers with JSON metadata and
never with a file. That is asserted on the content type, the body and the absence
of a `Content-Disposition` header — because the previous implementation answered
with bytes, and that is exactly why a click downloaded immediately.

THE LIBRARY IS REAL. A report exists after generation, survives a fresh
connection, and every row corresponds to stored artifacts. An empty library is
empty; there are no seeded rows.

READY MEANS READY. A row is never READY without bytes behind it, a download
button is offered only for a format that was actually written, and a delete takes
the metadata and the artifacts together.

THE KPI CHAIN HOLDS. `service.kpis` -> the stored preview -> the workbook -> the
PDF all carry the same figures at the same precision, for the same scope. This is
the test that would have caught the two defects this task fixed: PEI printed as
"66.00" where the card says "66", and Cannibalization's wider-scope fallback
dropped entirely.

SCOPE AND MODE DO NOT LEAK. Two scopes produce two different reports; the three
Simulation Studio modes produce three reports carrying only their own content.

Run with:

    ../venv/Scripts/python.exe -m pytest tests/test_report_center.py -q
"""

from __future__ import annotations

import io
import re

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader

from app.main import app
from app.store import reports as report_store
from app.store import db
from app.tpo import service as tpo_service
from app.tpo.filters import FilterState

client = TestClient(app)

URL = "/api/reports"

CC_SCOPE = {"year": 2025, "month": 10, "channel": ["CH002"]}
GO_SCOPE = {"month": 6, "category": ["Baby Care"], "channel": ["CH002"]}
TR_SCOPE = {"year": 2025, "month": 10, "channel": ["CH002"], "category": ["Baby Care"]}
SIM_OPTIONS = {"discount_pct": 15.0, "scenario_id": "optimized-plan",
               "scenario_name": "Optimized Plan"}
TR_OPTIONS = {"target_units": 50000.0, "current_discount_pct": 10.0, "checkpoint": 3}


@pytest.fixture(scope="module", autouse=True)
def isolated_store(tmp_path_factory):
    """A throwaway Report Center, the same `db.use_path` seam the store tests use."""
    path = tmp_path_factory.mktemp("report-center") / "reports.db"
    db.use_path(path)
    yield path
    db.close()


@pytest.fixture(autouse=True)
def empty_library():
    """Each test starts with an empty library, so counts mean what they say."""
    for row in report_store.listing(limit=500):
        report_store.delete(row.id)
    yield


def generate(module: str, scope: dict, options: dict | None = None,
             formats: list[str] | None = None, currency: str = "INR"):
    body = {"module": module, "scope": scope, "options": options or {}, "currency": currency}
    if formats is not None:
        body["formats"] = formats
    return client.post(URL, json=body)


def created(module: str, scope: dict, options: dict | None = None, **kw) -> dict:
    response = generate(module, scope, options, **kw)
    assert response.status_code == 201, response.text
    return response.json()


def pdf_text(payload: bytes) -> str:
    return "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(payload)).pages)


def artifact(report_id: str, fmt: str) -> tuple[bytes, str]:
    response = client.get(f"{URL}/{report_id}/download/{fmt}")
    assert response.status_code == 200, response.text
    name = re.search(r'filename="([^"]+)"', response.headers["content-disposition"]).group(1)
    return response.content, name


# --- generate is not download ------------------------------------------------


def test_generating_returns_metadata_and_never_a_file() -> None:
    """THE HEADLINE REQUIREMENT. Clicking Export Report must not hand the browser
    a file; it must create one in the Report Center."""
    response = generate("command-center", CC_SCOPE)
    assert response.status_code == 201
    assert response.headers["content-type"].startswith("application/json")
    # The three things a browser needs to start a download, all absent.
    assert "content-disposition" not in {k.lower() for k in response.headers}
    assert not response.content.startswith(b"PK\x03\x04")
    assert not response.content.startswith(b"%PDF-")

    body = response.json()
    assert body["status"] == "ready"
    assert body["report_id"]
    assert body["formats"]["xlsx"] and body["formats"]["pdf"]


def test_only_the_download_route_answers_with_a_file() -> None:
    report = created("command-center", CC_SCOPE)
    for fmt, magic in (("xlsx", b"PK\x03\x04"), ("pdf", b"%PDF-")):
        payload, name = artifact(report["report_id"], fmt)
        assert payload.startswith(magic)
        assert name.endswith(f".{fmt}")


def test_no_route_other_than_download_emits_an_attachment() -> None:
    """A sweep, so a future route cannot quietly reintroduce the old behaviour."""
    report = created("command-center", CC_SCOPE)
    for method, path in (
        ("post", URL),
        ("get", URL),
        ("get", f"{URL}/modules"),
        ("get", f"{URL}/{report['report_id']}"),
    ):
        call = client.post if method == "post" else client.get
        response = (
            call(path, json={"module": "command-center", "scope": CC_SCOPE})
            if method == "post" else call(path)
        )
        headers = {k.lower() for k in response.headers}
        assert "content-disposition" not in headers, f"{method} {path} answered with a file"


# --- the library is real -----------------------------------------------------


def test_an_empty_library_is_empty_and_says_so() -> None:
    """No seeded rows. The page this replaced shipped six authored ones."""
    body = client.get(URL).json()
    assert body["reports"] == []
    assert body["total"] == 0
    # The module list still travels, so the page can explain where reports come
    # from rather than showing a bare empty table.
    assert len(body["modules"]) == 5


def test_a_generated_report_appears_in_the_library() -> None:
    report = created("command-center", CC_SCOPE)
    body = client.get(URL).json()
    assert body["total"] == 1
    row = body["reports"][0]
    assert row["report_id"] == report["report_id"]
    assert row["module"] == "command-center"
    assert row["module_label"] == "Command Center"
    assert row["scope_label"] == "October F25 · Modern Trade"
    assert row["status"] == "ready"
    assert row["available_formats"] == ["pdf", "xlsx"]


def test_a_report_survives_a_fresh_connection() -> None:
    """Persistence, not a frontend list pretending to be saved. The connection is
    closed and reopened, so the row is read back off disk."""
    report = created("simulation-target-rescue", TR_SCOPE, TR_OPTIONS)
    db.close()
    row = report_store.get(report["report_id"])
    assert row.status == report_store.READY
    payload, _ = artifact(report["report_id"], "pdf")
    assert payload.startswith(b"%PDF-")


def test_the_library_stores_the_full_scope_and_filters() -> None:
    """Section 4's metadata: every selected dimension is recorded with the report,
    including the ones left unconstrained."""
    report = created("simulation-target-rescue", TR_SCOPE, TR_OPTIONS)
    row = client.get(f"{URL}/{report['report_id']}").json()
    assert row["scope"] == TR_SCOPE
    labels = {label for label, _ in row["filters"]}
    for expected in ("Year", "Month", "Channel", "Category", "Product", "Promotion"):
        assert expected in labels
    values = dict(row["filters"])
    assert values["Year"].startswith("2025")
    assert values["Channel"] == "Modern Trade"
    assert values["Category"] == "Baby Care"
    assert values["Product"] == "All"
    assert row["currency"] == "INR"
    assert row["created_at"]


def test_no_credential_shaped_option_is_ever_stored() -> None:
    """Section 4 and 30: a report is stored and read back later, so what goes into
    it is a data-safety decision."""
    report = created("simulation-target-rescue", TR_SCOPE, {
        **TR_OPTIONS, "auth_token": "sekrit", "password": "hunter2", "api_key": "abc",
    })
    row = report_store.get(report["report_id"])
    assert "target_units" in row.options
    for leaked in ("auth_token", "password", "api_key"):
        assert leaked not in row.options
    assert "sekrit" not in str(row.options)


# --- ready means ready -------------------------------------------------------


def test_a_single_format_report_offers_only_that_format() -> None:
    """A download button is never offered for a file that was not written."""
    report = created("command-center", CC_SCOPE, formats=["pdf"])
    assert report["formats"]["pdf"]
    assert report["formats"]["xlsx"] is None
    assert report["available_formats"] == ["pdf"]

    assert client.get(f"{URL}/{report['report_id']}/download/pdf").status_code == 200
    missing = client.get(f"{URL}/{report['report_id']}/download/xlsx")
    assert missing.status_code == 404
    assert "no XLSX artifact" in missing.json()["detail"]


def test_a_request_that_cannot_be_reported_on_leaves_no_ready_row() -> None:
    """Target Rescue without a target is a rejected REQUEST, not a failed report:
    the library must not fill with rows for calls that never should have run."""
    response = generate("simulation-target-rescue", TR_SCOPE, {"current_discount_pct": 10})
    assert response.status_code == 422
    assert "target" in response.json()["detail"].lower()
    assert client.get(URL).json()["total"] == 0


def test_a_badly_typed_scope_value_is_a_422_naming_the_key() -> None:
    """A malformed scope is a rejected REQUEST, not a server error.

    `year` and `month` are scalars on FilterState and JSON cannot say so, so
    `{"month": [6]}` used to build a state holding a list and blow up much later
    inside a cached lookup -- an unhashable-type TypeError surfacing as a 500 on
    what is plainly a bad request. Each case here names the offending key.
    """
    for scope, key in (
        ({**CC_SCOPE, "month": [6]}, "month"),
        ({**CC_SCOPE, "year": "2024"}, "year"),
        ({**CC_SCOPE, "month": 6.5}, "month"),
        ({**CC_SCOPE, "year": True}, "year"),
        ({**CC_SCOPE, "channel": "CH001"}, "channel"),
    ):
        response = generate("command-center", scope)
        assert response.status_code == 422, (scope, response.status_code)
        assert key in response.json()["detail"], (scope, response.json())

    # And nothing was stored for any of them.
    assert client.get(URL).json()["total"] == 0


def test_an_unknown_module_is_a_404_and_stores_nothing() -> None:
    response = generate("promotion-intelligence", CC_SCOPE)
    assert response.status_code == 404
    assert client.get(URL).json()["total"] == 0


def test_a_failed_report_is_never_shown_as_ready() -> None:
    """The store refuses to flip a row to READY without bytes."""
    report_id = report_store.begin(
        module="command-center", module_label="Command Center", name="n", title="t",
        scope_label="s", scope={}, options={}, currency="INR",
    )
    report_store.finish(report_id, artifacts={"xlsx": ("x.xlsx", b"")},
                        filters=[], preview={})
    row = report_store.get(report_id)
    assert row.status == report_store.FAILED
    assert row.error
    assert client.get(f"{URL}/{report_id}/download/xlsx").status_code == 404


def test_delete_removes_the_report_and_its_artifacts_together() -> None:
    report = created("command-center", CC_SCOPE)
    report_id = report["report_id"]
    assert client.get(f"{URL}/{report_id}/download/pdf").status_code == 200

    assert client.delete(f"{URL}/{report_id}").status_code == 204
    assert client.get(URL).json()["total"] == 0
    assert client.get(f"{URL}/{report_id}").status_code == 404
    # No orphan: the artifact goes with the row, in one statement.
    assert client.get(f"{URL}/{report_id}/download/pdf").status_code == 404
    assert client.delete(f"{URL}/{report_id}").status_code == 404


def test_clear_empties_the_whole_library_and_its_artifacts() -> None:
    """The Clear action. Every report goes, and every artifact with it."""
    first = created("command-center", CC_SCOPE)
    second = created("simulation-target-rescue", TR_SCOPE, TR_OPTIONS)
    assert client.get(URL).json()["total"] == 2

    response = client.delete(URL)
    assert response.status_code == 200
    assert response.json() == {"deleted": 2, "total": 0}

    body = client.get(URL).json()
    assert body["reports"] == []
    assert body["total"] == 0
    # No orphaned artifact survives its row.
    for report in (first, second):
        assert client.get(f"{URL}/{report['report_id']}").status_code == 404
        assert client.get(f"{URL}/{report['report_id']}/download/pdf").status_code == 404
        assert client.get(f"{URL}/{report['report_id']}/download/xlsx").status_code == 404


def test_clearing_an_empty_library_is_a_success_not_an_error() -> None:
    """Asking for a state you are already in is not a failure."""
    assert client.get(URL).json()["total"] == 0
    response = client.delete(URL)
    assert response.status_code == 200
    assert response.json() == {"deleted": 0, "total": 0}


def test_clear_is_not_limited_to_the_filtered_view() -> None:
    """A clear that spared what a filter was hiding would leave reports behind in
    a library the user believes is empty."""
    created("command-center", CC_SCOPE)
    created("simulation-target-rescue", TR_SCOPE, TR_OPTIONS)
    # The page may well be filtered to one module when Clear is pressed.
    filtered = client.get(URL, params={"module": "command-center"}).json()
    assert filtered["returned"] == 1
    assert filtered["total"] == 2

    assert client.delete(URL).json()["deleted"] == 2
    assert client.get(URL).json()["total"] == 0


def test_clear_frees_the_stored_bytes() -> None:
    """The artifacts are rows, so clearing genuinely releases them rather than
    leaving files behind that nothing points at."""
    created("command-center", CC_SCOPE)
    connection = db.connect()
    stored = connection.execute(
        "SELECT COALESCE(SUM(LENGTH(xlsx_blob)), 0) + COALESCE(SUM(LENGTH(pdf_blob)), 0) AS n"
        " FROM reports"
    ).fetchone()["n"]
    assert stored > 0

    client.delete(URL)
    after = connection.execute(
        "SELECT COALESCE(SUM(LENGTH(xlsx_blob)), 0) + COALESCE(SUM(LENGTH(pdf_blob)), 0) AS n"
        " FROM reports"
    ).fetchone()["n"]
    assert after == 0


# --- THE KPI CHAIN -----------------------------------------------------------


def test_ui_api_preview_excel_and_pdf_all_carry_the_same_kpis() -> None:
    """SECTION 34, end to end and at precision.

    `service.kpis` is what the Command Center screen renders from, so its
    `display_value` IS the number on the tile. That string must appear unchanged
    in the stored preview, in the workbook and in the PDF — for every one of the
    six cards.

    This is the test that catches precision drift. Before this task the PDF
    printed the card's "66" as "66.00", because the report re-rendered the raw
    value instead of carrying the card's own rendering.
    """
    state = FilterState.build(year=2025)
    cards = tpo_service.kpis(state, "INR")["kpis"]
    assert len(cards) == 6

    report = created("command-center", {"year": 2025})
    preview = {k["label"]: k for k in report["preview"]["kpis"]}

    workbook = load_workbook(io.BytesIO(artifact(report["report_id"], "xlsx")[0]))
    sheet = workbook["Executive Summary"]
    strings = {
        str(c.value) for row in sheet.iter_rows() for c in row if isinstance(c.value, str)
    }
    numbers = {
        round(float(c.value), 4)
        for row in sheet.iter_rows() for c in row if isinstance(c.value, (int, float))
    }
    document = pdf_text(artifact(report["report_id"], "pdf")[0])
    # The PDF substitutes the rupee sign it cannot draw; compare on that basis.
    document_normalised = document.replace("Rs.", "₹")

    checked = 0
    for card in cards.values():
        label = card["label"]
        assert label in preview, f"{label} missing from the stored preview"

        if card["value"] is None:
            # Not zero, and not silently dropped: the reason travels instead.
            assert not preview[label]["available"]
            assert preview[label]["display"]
            continue

        shown = card["display_value"]
        assert preview[label]["display"] == shown, f"{label}: preview disagrees with the API"
        assert shown in strings, f"{label}: {shown!r} is not in the workbook"
        assert shown in document_normalised, f"{label}: {shown!r} is not in the PDF"
        # And the workbook carries the RAW number too, so it can be summed.
        assert round(float(card["value"]), 4) in numbers, f"{label}: raw value missing"
        checked += 1

    assert checked >= 5, "too few KPIs were actually compared"


def test_the_pei_precision_defect_stays_fixed() -> None:
    """The brief's own worked example: the card shows 66 and -6.0%, so the report
    must show 66 and -6.0% — not 65.7 and not 66.00."""
    cards = tpo_service.kpis(FilterState.build(year=2025), "INR")["kpis"]
    pei = cards["pei"]
    assert pei["display_value"] == "66"

    report = created("command-center", {"year": 2025})
    document = pdf_text(artifact(report["report_id"], "pdf")[0])
    line = next(l for l in document.splitlines() if "Promotion Efficiency Index" in l)
    assert "66.00" not in document
    assert pei["delta_display"] in document
    assert line


def test_a_kpi_the_scope_cannot_support_still_shows_its_wider_measurement() -> None:
    """The Command Center tile falls back to the narrowest WIDER scope that can
    measure cannibalization, and names it. A report that dropped that showed less
    than the screen and read as a missing KPI — which is what was reported."""
    state = FilterState.build(year=2025, month=10, channel=["CH002"])
    card = tpo_service.kpis(state, "INR")["kpis"]["cannibalization_rate"]
    assert card["available"] is False
    wider = card["measured_at"]
    assert wider, "this scope no longer has a wider-scope fallback to check"

    report = created("command-center", CC_SCOPE)
    document = pdf_text(artifact(report["report_id"], "pdf")[0])
    assert wider["display_value"] in document
    assert wider["scope_label"] in document
    assert f"{wider['comparable_events']:,}" in document


# --- scope and mode do not leak ----------------------------------------------


def test_two_scopes_produce_two_different_reports() -> None:
    """Section 29: generate, change a filter, generate again."""
    first = created("command-center", CC_SCOPE)
    second = created("command-center", {**CC_SCOPE, "channel": ["CH003"]})

    assert first["report_id"] != second["report_id"]
    assert first["scope_label"] != second["scope_label"]
    assert "Modern Trade" in first["scope_label"]
    assert "General Trade" in second["scope_label"]

    def figures(report: dict) -> set[float]:
        book = load_workbook(io.BytesIO(artifact(report["report_id"], "xlsx")[0]))
        return {
            round(float(c.value), 2)
            for name in book.sheetnames
            for row in book[name].iter_rows()
            for c in row
            if isinstance(c.value, (int, float))
        }

    assert figures(first) != figures(second)
    assert client.get(URL).json()["total"] == 2


def test_the_three_simulation_modes_store_three_separate_reports() -> None:
    """Section 29 and 16: switch mode, generate, and carry nothing across."""
    investigation = created("simulation-investigation", CC_SCOPE, SIM_OPTIONS)
    optimization = created("simulation-general-optimization", GO_SCOPE, {})
    rescue = created("simulation-target-rescue", TR_SCOPE, TR_OPTIONS)

    library = client.get(URL).json()
    assert library["total"] == 3
    assert {r["module"] for r in library["reports"]} == {
        "simulation-investigation",
        "simulation-general-optimization",
        "simulation-target-rescue",
    }

    texts = {
        r["module"]: pdf_text(artifact(r["report_id"], "pdf")[0])
        for r in (investigation, optimization, rescue)
    }
    assert "Current Plan vs simulated scenario" in texts["simulation-investigation"]
    assert "Optimized product plan" in texts["simulation-general-optimization"]
    assert "Intervention comparison" in texts["simulation-target-rescue"]

    assert "Optimized product plan" not in texts["simulation-target-rescue"]
    assert "Intervention comparison" not in texts["simulation-general-optimization"]
    assert "Intervention comparison" not in texts["simulation-investigation"]


# --- listing, filtering, preview ---------------------------------------------


def test_the_library_filters_by_module_format_and_search() -> None:
    created("command-center", CC_SCOPE)
    created("simulation-target-rescue", TR_SCOPE, TR_OPTIONS, formats=["pdf"])

    by_module = client.get(URL, params={"module": "command-center"}).json()
    assert by_module["returned"] == 1
    assert by_module["reports"][0]["module"] == "command-center"

    by_format = client.get(URL, params={"format": "xlsx"}).json()
    assert by_format["returned"] == 1, "the PDF-only report should not match an xlsx filter"
    assert by_format["reports"][0]["module"] == "command-center"

    by_search = client.get(URL, params={"search": "target rescue"}).json()
    assert by_search["returned"] == 1
    assert by_search["reports"][0]["module"] == "simulation-target-rescue"

    assert client.get(URL, params={"search": "nothing matches this"}).json()["returned"] == 0


def test_the_library_is_newest_first() -> None:
    first = created("command-center", CC_SCOPE)
    second = created("command-center", {**CC_SCOPE, "channel": ["CH003"]})
    ids = [r["report_id"] for r in client.get(URL).json()["reports"]]
    assert ids[0] == second["report_id"]
    assert ids[1] == first["report_id"]


def test_the_preview_is_the_one_the_report_was_generated_with() -> None:
    """Section 9. Enough to confirm the report is the right one, and stored — not
    a fresh evaluation that could disagree with the artifacts beside it."""
    report = created("simulation-target-rescue", TR_SCOPE, TR_OPTIONS)
    preview = client.get(f"{URL}/{report['report_id']}").json()["preview"]

    assert preview["module"].endswith("Target Rescue")
    assert preview["title"] == "Monthly Target Recovery Report"
    assert preview["scope_line"] == report["scope_label"]
    assert preview["generated_display"]
    assert preview["headline"]
    assert preview["highlights"], "no summary lines captured"
    assert any("not a forecast" in d for d in preview["disclaimers"])

    # It is stored, so it does not move when the module would be re-run.
    again = client.get(f"{URL}/{report['report_id']}").json()["preview"]
    assert again == preview


def test_the_report_name_is_readable_not_a_filename() -> None:
    """Section 23: a person scanning a library reads a name, not a filename."""
    report = created("command-center", CC_SCOPE)
    assert report["name"] == "Command Center — October F25 · Modern Trade"
    assert ".xlsx" not in report["name"]
    assert "_" not in report["name"]
    # The FILENAME is still the sanitised one, on the artifact.
    assert report["formats"]["xlsx"] == "TPO_Command_Center_2025_Oct_Modern_Trade.xlsx"


def test_ownership_is_not_invented() -> None:
    """This project has no authentication, so a report has no verified author and
    none is manufactured — the same rule the rest of the store follows."""
    report = created("command-center", CC_SCOPE)
    assert report["owner"] is None
    assert "no authentication" in client.get(URL).json()["owner_note"]
