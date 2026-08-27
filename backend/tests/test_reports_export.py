"""Validation for report export — REAL FILES, OPENED AND READ BACK.

A 200 proves nothing about a report. Every test here generates bytes through the
endpoint, opens them with a reader that would reject a malformed file
(`openpyxl.load_workbook`, `pypdf.PdfReader`), and asserts on what is actually
inside: the sheets, the headers, the typed cells, the page count, the extracted
text, the scope line and the KPI values.

Four things are being defended.

THE FILES ARE REAL. A CSV renamed .xlsx does not load as a workbook, and HTML
renamed .pdf has no %PDF header and no page tree. Both are asserted directly.

THE NUMBERS ARE THE ENGINE'S. A KPI in a workbook is compared against
`service.kpis` for the same scope — the same function the Command Center screen
reads — so the export cannot have recomputed or rounded its way to a different
answer.

THE SCOPE PROPAGATES. Change the channel, the month, the category, the product or
the checkpoint, and the file changes with it. Two scopes never produce the same
bytes, and no report describes a scope it was not asked for.

THE THREE SIMULATION MODES ARE ISOLATED. A Target Rescue export carries no
optimizer product plan; a General Optimization export carries no intervention
ladder. Asserted on the extracted content, not on the request.

Run with:

    ../venv/Scripts/python.exe -m pytest tests/test_reports_export.py -q
"""

from __future__ import annotations

import io
import re

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader

from app.main import app
from app.reports import service as report_service
from app.store import reports as report_store
from app.store import db
from app.tpo import service as tpo_service
from app.tpo.filters import FilterState

client = TestClient(app)


@pytest.fixture(scope="module", autouse=True)
def isolated_store(tmp_path_factory):
    """A throwaway Report Center for this module.

    The same `db.use_path` seam `tests/test_store_persistence.py` uses. Without
    it these tests would fill the developer's real library with fixtures, and
    the count assertions would depend on whatever was already in it.
    """
    path = tmp_path_factory.mktemp("reports") / "reports.db"
    db.use_path(path)
    yield path
    db.close()

URL = "/api/reports"

XLSX_MAGIC = b"PK\x03\x04"          # a zip container; a CSV has no such header
PDF_MAGIC = b"%PDF-"

CC_SCOPE = {"year": 2025, "month": 10, "channel": ["CH002"]}
SIM_SCOPE = {"year": 2025, "month": 10, "channel": ["CH002"]}
GO_SCOPE = {"month": 6, "category": ["Baby Care"], "channel": ["CH002"]}
TR_SCOPE = {"year": 2025, "month": 10, "channel": ["CH002"], "category": ["Baby Care"]}

SIM_OPTIONS = {"discount_pct": 15.0, "scenario_id": "optimized-plan",
               "scenario_name": "Optimized Plan", "filename_hint": "Optimized Plan"}
GO_OPTIONS = {"min_discount_pct": 0.0, "max_discount_pct": 25.0}
TR_OPTIONS = {"target_units": 50000.0, "current_discount_pct": 10.0, "checkpoint": 3}

#: Every module that can be exported, with a scope and the control values its
#: authoritative service needs. Decision Center is exercised separately -- it is
#: assembled from posted Simulation Studio results rather than from a scope.
CASES: tuple[tuple[str, dict, dict], ...] = (
    ("command-center", CC_SCOPE, {}),
    ("simulation-investigation", SIM_SCOPE, SIM_OPTIONS),
    ("simulation-general-optimization", GO_SCOPE, GO_OPTIONS),
    ("simulation-target-rescue", TR_SCOPE, TR_OPTIONS),
)


def export(module: str, fmt: str, scope: dict, options: dict | None = None,
           currency: str = "INR"):
    """Generate one report into the Report Center.

    `fmt` selects which artifact to produce, so a test that only needs a PDF does
    not pay for a workbook.
    """
    return client.post(URL, json={
        "module": module, "scope": scope, "options": options or {},
        "currency": currency, "formats": [fmt],
    })


def ok(module: str, fmt: str, scope: dict, options: dict | None = None,
       currency: str = "INR") -> tuple[bytes, str]:
    """Generate, then DOWNLOAD — the two halves of the real workflow.

    Deliberately not a shortcut past the store: every byte these tests assert on
    has been written to the Report Center and read back out of it, which is what
    makes them evidence that a person clicking Excel gets that file.
    """
    created = export(module, fmt, scope, options, currency)
    assert created.status_code == 201, created.text
    record = created.json()
    assert record["status"] == "ready", record

    response = client.get(f"/api/reports/{record['report_id']}/download/{fmt}")
    assert response.status_code == 200, response.text
    name = re.search(r'filename="([^"]+)"', response.headers["content-disposition"]).group(1)
    return response.content, name


def pdf_text(payload: bytes) -> str:
    return "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(payload)).pages)


# --- the files are real ------------------------------------------------------


@pytest.mark.parametrize("module,scope,options", CASES)
def test_xlsx_is_a_real_workbook(module: str, scope: dict, options: dict) -> None:
    """Not a CSV with the extension changed: it opens as a workbook, it has the
    OOXML zip header, and it carries more than one worksheet's worth of
    structure."""
    payload, name = ok(module, "xlsx", scope, options)
    assert payload[:4] == XLSX_MAGIC
    assert name.endswith(".xlsx")
    assert len(payload) > 4000

    book = load_workbook(io.BytesIO(payload))
    assert book.sheetnames[0] == "Executive Summary"
    summary = book["Executive Summary"]
    values = [c.value for row in summary.iter_rows() for c in row if c.value is not None]
    assert "TPO INTELLIGENCE" in values
    # Every sheet name is legal and unique -- Excel silently corrupts a workbook
    # that breaks either rule.
    assert len(set(book.sheetnames)) == len(book.sheetnames)
    for sheet in book.sheetnames:
        assert len(sheet) <= 31
        assert not set(sheet) & set('[]:*?/\\')


@pytest.mark.parametrize("module,scope,options", CASES)
def test_pdf_is_a_real_pdf(module: str, scope: dict, options: dict) -> None:
    """Not HTML with the extension changed: it has the %PDF header, a readable
    page tree, document metadata and extractable text."""
    payload, name = ok(module, "pdf", scope, options)
    assert payload.startswith(PDF_MAGIC)
    assert b"%%EOF" in payload[-2048:]
    assert name.endswith(".pdf")

    reader = PdfReader(io.BytesIO(payload))
    assert len(reader.pages) >= 1
    assert reader.metadata is not None
    assert "TPO INTELLIGENCE" in (reader.metadata.title or "")

    text = pdf_text(payload)
    assert "TPO INTELLIGENCE" in text
    assert "Generated" in text
    # A page number on every page is what makes a printed extract navigable.
    assert "Page 1" in text


@pytest.mark.parametrize("module,scope,options", CASES)
def test_pdf_has_page_numbers_on_every_page(module: str, scope: dict, options: dict) -> None:
    payload, _ = ok(module, "pdf", scope, options)
    reader = PdfReader(io.BytesIO(payload))
    for n, page in enumerate(reader.pages, start=1):
        assert f"Page {n}" in (page.extract_text() or ""), f"page {n} has no page number"


def test_wide_tables_get_landscape_pages() -> None:
    """A wide table is re-framed, not clipped. The Command Center's alert table
    and the optimizer's product plan both ask for landscape, so those documents
    genuinely carry landscape pages beside their portrait ones."""
    for module, scope, options in (("command-center", CC_SCOPE, {}),
                                   ("simulation-general-optimization", GO_SCOPE, GO_OPTIONS)):
        payload, _ = ok(module, "pdf", scope, options)
        sizes = {
            (round(float(p.mediabox.width)), round(float(p.mediabox.height)))
            for p in PdfReader(io.BytesIO(payload)).pages
        }
        assert any(w > h for w, h in sizes), f"{module} has no landscape page"
        assert any(h > w for w, h in sizes), f"{module} has no portrait page"


# --- the numbers are the engine's -------------------------------------------


def test_command_center_kpis_match_the_engine() -> None:
    """The workbook's KPI values ARE `service.kpis`'s values for the same scope.

    This is the test that would fail if the export ever recalculated anything:
    the numbers are compared against the function the Command Center screen
    itself reads, not against a fixture.
    """
    state = FilterState.build(year=2025, month=10, channel=["CH002"])
    expected = tpo_service.kpis(state, "INR")["kpis"]

    payload, _ = ok("command-center", "xlsx", CC_SCOPE)
    summary = load_workbook(io.BytesIO(payload))["Executive Summary"]
    numeric = {
        round(float(c.value), 2)
        for row in summary.iter_rows() for c in row
        if isinstance(c.value, (int, float))
    }

    checked = 0
    for card in expected.values():
        if card.get("value") is None:
            continue
        assert round(float(card["value"]), 2) in numeric, f"{card['label']} is not in the workbook"
        checked += 1
    assert checked >= 4, "too few KPIs were actually compared"


def test_command_center_carries_its_promotion_performance_sections() -> None:
    """Brief section 6D. Promotion mix and promotion performance are part of the
    report, and a wrong payload key would silently drop them -- which is exactly
    what happened before this test existed."""
    payload, _ = ok("command-center", "xlsx", CC_SCOPE)
    book = load_workbook(io.BytesIO(payload))
    assert "Promotion Performance" in book.sheetnames

    performance = book["Promotion Performance"]
    headers = [c.value for row in performance.iter_rows(min_row=1, max_row=6) for c in row
               if isinstance(c.value, str)]
    for column in ("Promotion", "Product", "Channel", "Period", "Trade spend", "ROI", "Status"):
        assert column in headers, f"{column} is missing from the performance sheet"
    assert performance.max_row > 6, "the performance sheet has no rows"

    summary = " ".join(
        str(c.value) for row in book["Executive Summary"].iter_rows() for c in row
        if isinstance(c.value, str)
    )
    assert "Promotion mix" in summary


# --- the alert total is the alert population, not a page size ---------------
#
# THE DEFECT THIS SECTION EXISTS FOR. `adapters.command_center` printed
# "Total alerts" as `counts.get("total", len(rows))`. `service.risk_alerts`
# returns no "total" key, so the fallback ran on every export and `rows` is the
# alert list truncated to `alert_limit` -- the report published its own page
# size as the alert population, contradicting the Critical/High/Medium
# breakdown printed immediately above it.

#: A scope wide enough that the alert population exceeds the export's row cap,
#: so truncation is actually exercised. A narrow scope would pass the assertion
#: for the wrong reason -- with fewer alerts than the cap, a row count and the
#: population are the same number and the bug is invisible.
WIDE_CC_SCOPE = {"year": 2025}


def _risk_summary(text: str) -> dict[str, int]:
    """The Risk summary's four figures, read back out of an exported file."""
    out: dict[str, int] = {}
    for label in ("Critical", "High", "Medium", "Total alerts"):
        found = re.search(rf"{label}\s+([\d,]+)", text)
        if found:
            out[label] = int(found.group(1).replace(",", ""))
    return out


def _xlsx_summary_text(payload: bytes) -> str:
    sheet = load_workbook(io.BytesIO(payload))["Executive Summary"]
    lines = [
        " ".join(str(c.value) for c in row if c.value is not None)
        for row in sheet.iter_rows()
    ]
    return chr(10).join(lines)


#: The severity labels the alert sheet's data rows carry. The sheet also holds
#: the report title, the table title and a header row; keying off the severity
#: column is what separates a real alert from that chrome.
_SEVERITIES = ("Critical", "High", "Medium")


def _alert_rows(payload: bytes) -> list[tuple[str, ...]]:
    """The alert sheet's DATA rows, identified by promotion event.

    Severity, promotion, product, channel and period together identify the
    event the alert was raised for -- the same grain `service.risk_alerts`
    bands. Trailing measure columns are excluded so two genuinely distinct
    events cannot be told apart by a rounding difference alone.
    """
    sheet = load_workbook(io.BytesIO(payload))["Risk Alerts"]
    rows = []
    for row in sheet.iter_rows():
        cells = [c.value for c in row]
        if cells and str(cells[0]).strip() in _SEVERITIES:
            rows.append(tuple(str(c) for c in cells[:5]))
    return rows


@pytest.mark.parametrize("scope", [CC_SCOPE, WIDE_CC_SCOPE])
def test_exported_total_alerts_is_the_authoritative_alert_count(scope: dict) -> None:
    """A + B. The total IS the engine's banded population, and reconciles.

    Compared against `tpo_service.risk_alerts` -- the same function
    /api/command-center/risk-alerts calls -- for the same scope, so the export
    cannot have counted its way to a different answer. The bands are mutually
    exclusive (`service._severity` returns exactly one of them per event), so
    their sum is the alert count.
    """
    state = FilterState.build(**scope)
    counts = tpo_service.risk_alerts(state, "INR")["counts"]
    expected = counts["critical"] + counts["high"] + counts["medium"]

    summary = _risk_summary(_xlsx_summary_text(ok("command-center", "xlsx", scope)[0]))
    assert summary["Critical"] == counts["critical"]
    assert summary["High"] == counts["high"]
    assert summary["Medium"] == counts["medium"]
    assert summary["Total alerts"] == expected
    assert summary["Critical"] + summary["High"] + summary["Medium"] == summary["Total alerts"]


def test_exported_total_alerts_is_not_the_truncated_row_count() -> None:
    """The regression itself, stated directly.

    Under the wide scope the alert population is larger than the export's row
    cap, so the old `len(rows)` fallback would report exactly `alert_limit`.
    """
    state = FilterState.build(**WIDE_CC_SCOPE)
    counts = tpo_service.risk_alerts(state, "INR")["counts"]
    population = counts["critical"] + counts["high"] + counts["medium"]

    payload, _ = ok("command-center", "xlsx", WIDE_CC_SCOPE)
    listed = len(_alert_rows(payload))
    summary = _risk_summary(_xlsx_summary_text(payload))

    assert listed < population, "this scope must truncate, or the test proves nothing"
    assert summary["Total alerts"] == population
    assert summary["Total alerts"] != listed


def test_exported_total_alerts_is_not_the_promotion_event_count() -> None:
    """Nor the other number sitting in the same payload.

    `total_events` counts every promotion event in scope, on-target ones
    included. It is the denominator the screen reports as "N of M at target",
    and reporting it as the alert total would overstate the alerts.
    """
    state = FilterState.build(**WIDE_CC_SCOPE)
    counts = tpo_service.risk_alerts(state, "INR")["counts"]
    summary = _risk_summary(_xlsx_summary_text(ok("command-center", "xlsx", WIDE_CC_SCOPE)[0]))

    assert counts["total_events"] > counts["critical"] + counts["high"] + counts["medium"], (
        "this scope must contain on-target events, or the test proves nothing"
    )
    assert summary["Total alerts"] != counts["total_events"]


def test_the_total_alerts_figure_follows_the_filter_scope() -> None:
    """C. Two scopes, two populations, and each file carries its own."""
    narrow = {"year": 2025, "month": 10, "channel": ["CH002"]}
    wide = WIDE_CC_SCOPE

    def total_for(scope: dict) -> tuple[int, int]:
        counts = tpo_service.risk_alerts(FilterState.build(**scope), "INR")["counts"]
        engine = counts["critical"] + counts["high"] + counts["medium"]
        exported = _risk_summary(_xlsx_summary_text(ok("command-center", "xlsx", scope)[0]))
        return engine, exported["Total alerts"]

    narrow_engine, narrow_exported = total_for(narrow)
    wide_engine, wide_exported = total_for(wide)

    assert narrow_exported == narrow_engine
    assert wide_exported == wide_engine
    assert narrow_exported != wide_exported, "the two scopes must differ, or this proves nothing"


def test_both_formats_receive_the_same_total_alerts() -> None:
    """D + E. One value in the ReportDoc, so Excel and PDF cannot disagree."""
    state = FilterState.build(**WIDE_CC_SCOPE)
    counts = tpo_service.risk_alerts(state, "INR")["counts"]
    expected = counts["critical"] + counts["high"] + counts["medium"]

    from_xlsx = _risk_summary(_xlsx_summary_text(ok("command-center", "xlsx", WIDE_CC_SCOPE)[0]))
    from_pdf = _risk_summary(pdf_text(ok("command-center", "pdf", WIDE_CC_SCOPE)[0]))

    assert from_xlsx["Total alerts"] == expected
    assert from_pdf["Total alerts"] == expected
    assert from_xlsx == from_pdf


def test_the_alert_listing_introduces_no_duplicates() -> None:
    """F. No alert is counted twice, and the export adds no row of its own.

    Two things, because the total and the listing can fail independently.

    THE POPULATION IS DISTINCT. `event.key` is the promotion event's identity
    (product | channel | week | promotion), and the total is a sum of bands, so
    one event landing in the listing twice would inflate the figure the summary
    prints. Asserted on the engine's own ids.

    THE EXPORT ADDS NOTHING. The alert sheet carries exactly as many data rows
    as the service returned for the same cap -- no row is repeated on its way
    into the workbook, and none is dropped.

    Deliberately NOT asserted on the exported rows' own contents: the sheet's
    Period column is empty (`_alert_row` reads a key `risk_alerts` does not
    emit), so two events that differ only by week are indistinguishable there.
    That is a separate defect in the alert LISTING, outside this fix, and it
    must not be papered over by a test that quietly tolerates it.
    """
    state = FilterState.build(**WIDE_CC_SCOPE)
    served = tpo_service.risk_alerts(state, "INR", limit=200)["alerts"]
    ids = [a["id"] for a in served]
    assert ids, "the engine returned no alerts for this scope"
    assert len(ids) == len(set(ids)), "the alert population repeats an event"

    rows = _alert_rows(ok("command-center", "xlsx", WIDE_CC_SCOPE)[0])
    assert len(rows) == len(served), "the workbook and the engine disagree on row count"


def test_currency_cells_are_typed_numbers_not_text() -> None:
    """A currency column holds a NUMBER with a currency format, so the recipient
    can sum and sort it. Text like "₹90.7 L" would be unusable."""
    payload, _ = ok("command-center", "xlsx", CC_SCOPE)
    summary = load_workbook(io.BytesIO(payload))["Executive Summary"]
    money = [
        c for row in summary.iter_rows() for c in row
        if isinstance(c.value, (int, float)) and "₹" in (c.number_format or "")
    ]
    assert money, "no currency-formatted numeric cell found"
    assert all(isinstance(c.value, (int, float)) for c in money)


def test_usd_selection_is_respected_and_never_hard_coded() -> None:
    """The workbook follows the SELECTED currency. A USD session must not produce
    a rupee-formatted book."""
    inr, _ = ok("command-center", "xlsx", CC_SCOPE, currency="INR")
    usd, _ = ok("command-center", "xlsx", CC_SCOPE, currency="USD")

    def formats(payload: bytes) -> set[str]:
        sheet = load_workbook(io.BytesIO(payload))["Executive Summary"]
        return {
            c.number_format for row in sheet.iter_rows() for c in row
            if isinstance(c.value, (int, float)) and c.number_format
        }

    assert any("₹" in f for f in formats(inr))
    assert any("$" in f for f in formats(usd))
    assert not any("₹" in f for f in formats(usd)), "a USD export carried a rupee format"


def test_no_pdf_prints_an_undrawable_character() -> None:
    """A glyph the font cannot draw prints as a hollow box, which is worse than
    useless on a printed report. INR is therefore rendered "Rs." in PDFs -- the
    workbook keeps the real symbol, which Excel can draw."""
    for module, scope, options in CASES:
        text = pdf_text(ok(module, "pdf", scope, options)[0])
        assert "₹" not in text, f"{module} printed a rupee sign the font cannot draw"

    money = pdf_text(ok("simulation-target-rescue", "pdf", TR_SCOPE, TR_OPTIONS)[0])
    assert "Rs." in money, "no INR figure reached the PDF at all"


def test_a_usd_pdf_uses_the_dollar_sign() -> None:
    text = pdf_text(ok("command-center", "pdf", CC_SCOPE, {}, currency="USD")[0])
    assert "$" in text
    assert "Rs." not in text, "a USD report carried an INR figure"


def test_a_missing_kpi_is_absent_not_zero() -> None:
    """A KPI the engine could not produce must not appear as 0. The workbook
    carries its reason instead."""
    state = FilterState.build(year=2025, month=10, channel=["CH002"])
    cards = tpo_service.kpis(state, "INR")["kpis"]
    unavailable = [c for c in cards.values() if c.get("value") is None]
    if not unavailable:
        pytest.skip("every KPI is available for this scope")

    payload, _ = ok("command-center", "xlsx", CC_SCOPE)
    summary = load_workbook(io.BytesIO(payload))["Executive Summary"]
    text = " ".join(
        str(c.value) for row in summary.iter_rows() for c in row if isinstance(c.value, str)
    )
    for card in unavailable:
        assert card["label"] in text
        assert (card.get("unavailable_reason") or "Not available")[:30] in text


# --- the scope propagates ----------------------------------------------------


def test_the_scope_line_names_the_selection() -> None:
    payload, _ = ok("command-center", "pdf", CC_SCOPE)
    text = pdf_text(payload)
    assert "October F25" in text
    assert "Modern Trade" in text


def test_every_filter_dimension_is_named_even_when_unconstrained() -> None:
    """A Filters block that lists only what was set leaves the reader guessing
    whether Region was filtered or forgotten."""
    payload, _ = ok("command-center", "pdf", CC_SCOPE)
    text = pdf_text(payload)
    for label in ("Channel", "Region", "Retailer", "Category", "Product", "Promotion"):
        assert label in text, f"{label} is not named in the Filters block"
    assert "All" in text


@pytest.mark.parametrize("changed,key", [
    ({"channel": ["CH003"]}, "channel"),
    ({"month": 6}, "month"),
    ({"year": 2024}, "year"),
])
def test_changing_a_filter_changes_the_report(changed: dict, key: str) -> None:
    """Brief section 22. Every export reflects the CURRENT state; nothing is
    cached."""
    base, _ = ok("command-center", "xlsx", CC_SCOPE)
    other, _ = ok("command-center", "xlsx", {**CC_SCOPE, **changed})
    assert base != other, f"changing {key} produced an identical file"

    base_values = _numeric(base)
    other_values = _numeric(other)
    assert base_values != other_values, f"changing {key} did not change any figure"


def _numeric(payload: bytes) -> set[float]:
    """Every typed number in the WHOLE workbook.

    All sheets, not just the summary: a module's figures may live on a detail
    sheet, and a helper that only read the first one would silently compare two
    empty sets and pass.
    """
    book = load_workbook(io.BytesIO(payload))
    values = {
        round(float(c.value), 2)
        for name in book.sheetnames
        for row in book[name].iter_rows()
        for c in row
        if isinstance(c.value, (int, float))
    }
    assert values, "the workbook carries no typed numeric cell at all"
    return values


@pytest.mark.parametrize("changed", [
    {"category": ["Health Care"]},
    {"checkpoint_change": True},
])
def test_target_rescue_export_follows_its_own_controls(changed: dict) -> None:
    """Category, product and checkpoint all reach the file."""
    scope = dict(TR_SCOPE)
    options = dict(TR_OPTIONS)
    if changed.get("checkpoint_change"):
        options["checkpoint"] = 1
    else:
        scope.update(changed)

    base, _ = ok("simulation-target-rescue", "xlsx", TR_SCOPE, TR_OPTIONS)
    other, _ = ok("simulation-target-rescue", "xlsx", scope, options)
    assert _numeric(base) != _numeric(other)


def test_target_rescue_product_filter_reaches_the_file() -> None:
    from app.tpo.loader import get_store

    store = get_store()
    product = next(p for p, m in sorted(store.dims.products.items()) if m.category == "Baby Care")
    payload, _ = ok("simulation-target-rescue", "pdf",
                    {**TR_SCOPE, "product": [product]}, TR_OPTIONS)
    text = pdf_text(payload)
    assert store.dims.products[product].name.strip()[:18] in text


# --- the three simulation modes are isolated --------------------------------


def test_each_simulation_mode_exports_only_its_own_content() -> None:
    """Brief section 21. Switching mode and exporting again must not carry the
    previous mode's data across."""
    rescue_text = pdf_text(ok("simulation-target-rescue", "pdf", TR_SCOPE, TR_OPTIONS)[0])
    optimize_text = pdf_text(
        ok("simulation-general-optimization", "pdf", GO_SCOPE, GO_OPTIONS)[0])
    investigate_text = pdf_text(
        ok("simulation-investigation", "pdf", SIM_SCOPE, SIM_OPTIONS)[0])

    assert "Target Rescue" in rescue_text
    assert "Intervention comparison" in rescue_text
    assert "Optimized product plan" not in rescue_text
    assert "Trade Spend Optimization Report" not in rescue_text

    assert "General Optimization" in optimize_text
    assert "Optimized product plan" in optimize_text
    assert "Intervention comparison" not in optimize_text
    assert "Run-rate" not in optimize_text

    assert "Investigation Simulation" in investigate_text
    assert "Current Plan vs simulated scenario" in investigate_text
    assert "Optimized product plan" not in investigate_text
    assert "Intervention comparison" not in investigate_text


def test_each_mode_gets_its_own_worksheets() -> None:
    sheets = {}
    for module, scope, options in (
        ("simulation-investigation", SIM_SCOPE, SIM_OPTIONS),
        ("simulation-general-optimization", GO_SCOPE, GO_OPTIONS),
        ("simulation-target-rescue", TR_SCOPE, TR_OPTIONS),
    ):
        payload, _ = ok(module, "xlsx", scope, options)
        sheets[module] = set(load_workbook(io.BytesIO(payload)).sheetnames)

    assert "Current vs Simulated" in sheets["simulation-investigation"]
    assert "Optimized Plan" in sheets["simulation-general-optimization"]
    assert "Interventions" in sheets["simulation-target-rescue"]
    assert "Optimized Plan" not in sheets["simulation-target-rescue"]
    assert "Interventions" not in sheets["simulation-general-optimization"]


# --- module-specific content --------------------------------------------------


def test_target_rescue_separates_the_two_clocks() -> None:
    """Brief section 12C. The business-week coverage and the calendar month are
    reported separately and each is labelled, so a 28-day analytical month can
    never be read as a calendar day count."""
    text = pdf_text(ok("simulation-target-rescue", "pdf", TR_SCOPE, TR_OPTIONS)[0])
    assert "Analytical checkpoint" in text
    assert "completed business weeks" in text
    assert "Business-week coverage" in text
    assert "not calendar days" in text
    assert "Calendar month length" in text
    assert "calendar days" in text


def test_target_rescue_labels_the_run_rate_honestly() -> None:
    text = pdf_text(ok("simulation-target-rescue", "pdf", TR_SCOPE, TR_OPTIONS)[0])
    assert "Run-rate projection is a planning indicator, not a forecast." in text


def test_simulation_never_labels_a_scenario_as_actuals() -> None:
    for module, scope, options in (
        ("simulation-investigation", SIM_SCOPE, SIM_OPTIONS),
        ("simulation-general-optimization", GO_SCOPE, GO_OPTIONS),
        ("simulation-target-rescue", TR_SCOPE, TR_OPTIONS),
    ):
        text = pdf_text(ok(module, "pdf", scope, options)[0])
        assert "are not historical actuals" in text, module


def test_investigation_simulation_separates_measured_from_simulated() -> None:
    payload, _ = ok("simulation-investigation", "xlsx", SIM_SCOPE, SIM_OPTIONS)
    sheet = load_workbook(io.BytesIO(payload))["Current vs Simulated"]
    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=6)) if c.value]
    joined = " ".join(
        str(c.value) for row in sheet.iter_rows() for c in row if isinstance(c.value, str)
    )
    assert "Measured (Current Plan)" in joined
    assert "Simulated — low" in joined
    assert "Simulated — high" in joined
    assert headers is not None


def test_a_scenario_that_could_not_be_applied_says_so() -> None:
    """A column of 0.00 must not read as "the treatment earns nothing" when the
    truth is "the treatment could not be applied here". `execution.simulate`
    reports how many promoted rows it had to exclude, and the report carries that
    count and its reason next to the figures it explains."""
    text = pdf_text(ok("simulation-investigation", "pdf", SIM_SCOPE, SIM_OPTIONS)[0])
    assert "Scenario coverage" in text
    assert "Rows the treatment could not be" in text
    assert "cannot be re-based" in text


def test_the_pdf_kpi_table_shows_the_previous_period() -> None:
    """Brief section 6B wants the previous-period value beside each KPI. The
    workbook had it and the PDF was printing a dash, which made the two formats
    disagree about the same card."""
    text = pdf_text(ok("command-center", "pdf", CC_SCOPE)[0])
    state = FilterState.build(year=2025, month=10, channel=["CH002"])
    cards = tpo_service.kpis(state, "INR")["kpis"]
    with_previous = [c for c in cards.values() if c.get("previous_value") is not None]
    assert with_previous, "no KPI in this scope has a previous period to show"
    # The delta basis the cards use travels too, so the comparison is explained.
    assert "Previous" in text
    assert any(c.get("delta_display", "") in text for c in with_previous)


def test_no_report_claims_realtime_or_ai() -> None:
    """The brief forbids both claims unless they are true here, and neither is."""
    for module, scope, options in CASES:
        for fmt in ("pdf",):
            text = pdf_text(ok(module, fmt, scope, options)[0]).lower()
            assert "real-time" not in text, module
            assert "realtime" not in text, module
            assert "ai generated" not in text, module
            assert "ai-generated" not in text, module


def test_reports_carry_their_provenance() -> None:
    for module, scope, options in CASES:
        text = pdf_text(ok(module, "pdf", scope, options)[0])
        assert "Generated from the selected TPO Intelligence view" in text, module
        assert "Source" in text, module


# --- filenames ----------------------------------------------------------------


def test_filenames_are_predictable_and_sanitised() -> None:
    _, cc = ok("command-center", "xlsx", CC_SCOPE)
    assert cc == "TPO_Command_Center_2025_Oct_Modern_Trade.xlsx"

    _, tr = ok("simulation-target-rescue", "pdf", TR_SCOPE, TR_OPTIONS)
    assert tr == "TPO_Simulation_Target_Rescue_2025_Oct_Modern_Trade.pdf"

    _, go = ok("simulation-general-optimization", "xlsx", GO_SCOPE, GO_OPTIONS)
    assert go == "TPO_Simulation_General_Optimization_Jun_Modern_Trade.xlsx"

    for name in (cc, tr, go):
        assert not set(name) & set('<>:"/\\|?*')


def test_two_channels_do_not_collide_on_one_filename() -> None:
    """Two exports of the same month for different channels must not land on the
    same name -- that is the moment a reader loses track of which file is which."""
    _, mt = ok("command-center", "pdf", {"year": 2025, "month": 10, "channel": ["CH002"]})
    _, ec = ok("command-center", "pdf", {"year": 2025, "month": 10, "channel": ["CH001"]})
    _, both = ok("command-center", "pdf", {"year": 2025, "month": 10})
    assert len({mt, ec, both}) == 3
    assert "Modern_Trade" in mt and "E-commerce" in ec


def test_a_filename_hint_is_sanitised_and_carries_no_identifier() -> None:
    _, name = ok("simulation-investigation", "pdf", SIM_SCOPE, {
        **SIM_OPTIONS,
        "filename_hint": 'Diwali/Deal "25" *risky* 3f2504e0-4f89-11d3-9a0c-0305e82c3301',
    })
    assert not set(name) & set('<>:"/\\|?*')
    assert "3f2504e0" not in name, "a raw identifier reached the filename"
    assert name.endswith(".pdf")


def test_sanitize_strips_identifiers_and_unsafe_characters() -> None:
    assert "/" not in report_service.sanitize("a/b")
    assert report_service.sanitize("3f2504e0-4f89-11d3-9a0c-0305e82c3301") == ""
    assert report_service.sanitize("Dussehra Deal 25") == "Dussehra_Deal_25"


# --- errors are real ----------------------------------------------------------


def test_an_unknown_module_is_a_404_naming_the_supported_ones() -> None:
    response = export("promotion-intelligence", "pdf", CC_SCOPE)
    assert response.status_code == 404
    assert "command-center" in response.json()["detail"]


def test_an_unsupported_format_is_rejected() -> None:
    for bad in ("docx", "csv", "html"):
        assert client.post(URL, json={
            "module": "command-center", "scope": CC_SCOPE, "formats": [bad],
        }).status_code == 422, bad
    for bad in ("docx", "csv"):
        assert client.get(
            f"/api/reports/{'0' * 32}/download/{bad}"
        ).status_code == 422, bad


def test_target_rescue_without_a_target_is_a_422_with_the_reason() -> None:
    """Never a blank-looking report. The reader is told what to do."""
    response = export("simulation-target-rescue", "xlsx", TR_SCOPE, {"current_discount_pct": 10})
    assert response.status_code == 422
    assert "target" in response.json()["detail"].lower()


def test_decision_center_without_a_record_is_a_422() -> None:
    response = export("decision-center", "pdf", CC_SCOPE, {})
    assert response.status_code == 422
    detail = response.json()["detail"]
    assert "decision record" in detail.lower()
    assert "context" in detail


def test_an_unknown_filter_dimension_is_rejected() -> None:
    """`FilterState.build` owns this, so a client cannot smuggle a filter the
    project has no meaning for into a report."""
    assert export("command-center", "xlsx", {"not_a_dimension": ["x"]}).status_code == 422


def test_a_scope_with_no_rows_reports_the_reason_rather_than_an_empty_grid() -> None:
    payload, _ = ok("simulation-general-optimization", "pdf",
                    {"month": 2, "category": ["Baby Care"], "channel": ["CH002"]},
                    {"min_discount_pct": 24.0, "max_discount_pct": 24.0})
    text = pdf_text(payload)
    # A window with no approved treatment in it is a real answer, and the report
    # states it instead of printing a plan of zeros.
    assert "No plan" in text or "no approved" in text.lower()


# --- nothing internal leaks ---------------------------------------------------


def test_no_debug_or_internal_fields_reach_a_report() -> None:
    """Brief section 30. The engine's debug block, row indices and internal keys
    stay out of a business deliverable."""
    for module, scope, options in CASES:
        text = pdf_text(ok(module, "pdf", scope, options)[0]).lower()
        for forbidden in ("debug", "traceback", "password", "token", "secret",
                          "authorization", "comparable_events", "baseline_key"):
            assert forbidden not in text, f"{module} leaked {forbidden!r}"


def test_the_registry_only_carries_modules_with_a_computed_source() -> None:
    """Administrative screens are deliberately absent: the brief rules out export
    controls with no reportable dataset behind them."""
    keys = set(report_service.module_keys())
    assert keys == {
        "command-center", "simulation-investigation",
        "simulation-general-optimization", "simulation-target-rescue",
        "decision-center",
    }
    for absent in ("settings", "connections", "reports", "investigations",
                   "promotion-intelligence"):
        assert absent not in keys


def test_the_modules_endpoint_matches_the_registry() -> None:
    body = client.get("/api/reports/modules").json()
    assert {m["key"] for m in body["modules"]} == set(report_service.module_keys())
    assert set(body["formats"]) == {"xlsx", "pdf"}


# --- no existing endpoint changed --------------------------------------------


def test_export_does_not_disturb_the_endpoints_it_reads() -> None:
    """The report service calls the same functions the screens call. Running an
    export must leave those endpoints returning exactly what they returned
    before -- no cache poisoned, no state mutated."""
    params = {"year": 2025, "month": 10, "channel": ["CH002"]}
    before_kpis = client.get("/api/command-center/kpis", params=params).json()
    before_alerts = client.get("/api/command-center/risk-alerts", params=params).json()

    ok("command-center", "xlsx", CC_SCOPE)
    ok("command-center", "pdf", CC_SCOPE)
    ok("simulation-target-rescue", "pdf", TR_SCOPE, TR_OPTIONS)

    assert client.get("/api/command-center/kpis", params=params).json() == before_kpis
    assert client.get("/api/command-center/risk-alerts", params=params).json() == before_alerts


def test_the_report_layer_computes_no_business_figure() -> None:
    """A source-level guard on the architectural rule: the adapters may read the
    services and copy figures, but they may not import the KPI engine and derive
    one of their own."""
    import ast
    import inspect

    from app.reports import adapters

    tree = ast.parse(inspect.getsource(adapters))
    imported: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.ImportFrom) and node.module:
            imported.add(node.module.split(".")[-1])
            imported.update(a.name.split(".")[-1] for a in node.names)
        elif isinstance(node, ast.Import):
            imported.update(a.name.split(".")[-1] for a in node.names)

    # The KPI engine itself is NOT imported here: an adapter that could call
    # `aggregate.calculate_roi` could produce a second ROI.
    assert "aggregate" not in imported
    # What it does import are the module services the screens themselves call.
    assert {"service", "optimization", "rescue", "execution", "simulation"} <= imported
