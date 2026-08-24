"""The Excel writer — one `ReportDoc` in, real `.xlsx` bytes out.

A REAL WORKBOOK, via openpyxl. Not a CSV with the extension changed: what comes
out of here is the zipped OOXML a spreadsheet actually opens, with typed cells,
number formats, frozen headers and autofilters.

IT KNOWS NOTHING ABOUT THE BUSINESS. No KPI is named here, no module is special
cased, nothing is computed. It walks the sections an adapter built and lays them
out. That is what keeps fourteen format-times-module combinations down to two
writers.

NUMBERS GO IN AS NUMBERS. A currency cell receives `9071892.0` and a number
format, never the string "₹90.7 L" — so the recipient can sum a column, sort it
and chart it. The currency SYMBOL in that format follows the currency the user
had selected, so a USD session never yields a rupee-formatted book.

A MISSING VALUE STAYS EMPTY. `None` is written as a blank cell, never as 0: a
figure the engine could not produce must read as absent, exactly as it does on
screen.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.reports.model import ReportDoc, Section, Table, excel_number_format, is_numeric

# --- the one palette, matching the application's own chrome -----------------

_INK = "1A1F2E"
_MUTED = "6B7280"
_BRAND = "6B47FF"
_HEADER_BG = "F1F2F7"
_BAND = "FAFAFC"
_RULE = "E2E5EC"

_TITLE = Font(name="Calibri", size=18, bold=True, color=_INK)
_BRANDLINE = Font(name="Calibri", size=10, bold=True, color=_BRAND)
_H2 = Font(name="Calibri", size=12, bold=True, color=_INK)
_LABEL = Font(name="Calibri", size=10, color=_MUTED)
_VALUE = Font(name="Calibri", size=10, color=_INK)
_VALUE_BOLD = Font(name="Calibri", size=10, bold=True, color=_INK)
_TH = Font(name="Calibri", size=10, bold=True, color=_INK)
_SMALL = Font(name="Calibri", size=9, color=_MUTED)

_TH_FILL = PatternFill("solid", fgColor=_HEADER_BG)
_BAND_FILL = PatternFill("solid", fgColor=_BAND)
_THIN = Side(style="thin", color=_RULE)
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WRAP = Alignment(vertical="top", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _sheet_name(raw: str, used: set[str]) -> str:
    """A legal, unique worksheet name.

    Excel rejects []:*?/\\ and anything over 31 characters, and silently
    corrupts a workbook with duplicates — so both are handled here rather than
    left for an adapter to remember.
    """
    clean = "".join(c for c in raw if c not in set('[]:*?/\\')).strip() or "Sheet"
    clean = clean[:31]
    if clean not in used:
        used.add(clean)
        return clean
    for n in range(2, 100):
        suffix = f" {n}"
        candidate = clean[: 31 - len(suffix)] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
    used.add(clean[:28] + "~99")
    return clean[:28] + "~99"


def _put(ws: Worksheet, row: int, col: int, value: Any, *, font=None, fmt=None, align=None,
         fill=None, border=None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if fill:
        cell.fill = fill
    if border:
        cell.border = border


def _cover(ws: Worksheet, doc: ReportDoc, currency: str) -> int:
    """The Executive Summary block: brand, title, when, scope, filters, meta."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 62

    r = 1
    _put(ws, r, 1, doc.brand, font=_BRANDLINE); r += 1
    _put(ws, r, 1, doc.module, font=_H2); r += 1
    _put(ws, r, 1, doc.title, font=_TITLE); r += 2

    for label, value in (
        ("Generated", doc.generated_display),
        ("Scope", doc.scope_line),
    ):
        _put(ws, r, 1, label, font=_LABEL)
        _put(ws, r, 2, value, font=_VALUE_BOLD, align=_WRAP)
        r += 1

    if doc.headline:
        r += 1
        _put(ws, r, 1, "Status", font=_LABEL)
        _put(ws, r, 2, doc.headline, font=_VALUE_BOLD, align=_WRAP)
        r += 1

    if doc.empty_reason:
        r += 1
        _put(ws, r, 1, "No data", font=_LABEL)
        _put(ws, r, 2, doc.empty_reason, font=_VALUE, align=_WRAP)
        r += 1

    if doc.filters:
        r += 1
        _put(ws, r, 1, "Filters", font=_H2); r += 1
        for label, value in doc.filters:
            _put(ws, r, 1, label, font=_LABEL)
            _put(ws, r, 2, value, font=_VALUE, align=_WRAP)
            r += 1

    if doc.meta:
        r += 1
        _put(ws, r, 1, "Report metadata", font=_H2); r += 1
        for label, value in doc.meta:
            _put(ws, r, 1, label, font=_LABEL)
            _put(ws, r, 2, value, font=_VALUE, align=_WRAP)
            r += 1

    if doc.disclaimers:
        r += 1
        _put(ws, r, 1, "Notes", font=_H2); r += 1
        for line in doc.disclaimers:
            _put(ws, r, 1, line, font=_SMALL, align=_WRAP)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.row_dimensions[r].height = 28
            r += 1
    return r + 1


def _kpi_block(ws: Worksheet, row: int, section: Section, currency: str) -> int:
    """KPI cards as a small typed grid — value, previous, delta, trend."""
    _put(ws, row, 1, section.title, font=_H2)
    row += 1

    headers = ("KPI", "Value", "As displayed", "Previous", "Delta", "Trend",
               "Basis / evidence")
    for i, h in enumerate(headers, start=1):
        _put(ws, row, i, h, font=_TH, fill=_TH_FILL, border=_BOX, align=_LEFT)
    header_row = row
    row += 1

    for entry in section.items:
        _put(ws, row, 1, entry.label, font=_VALUE_BOLD, border=_BOX)
        if entry.available and entry.value is not None:
            _put(ws, row, 2, entry.value, font=_VALUE, border=_BOX, align=_RIGHT,
                 fmt=excel_number_format(entry.kind, currency))
        else:
            # NOT ZERO. The card had no value, and the reason travels with it.
            _put(ws, row, 2, entry.unavailable_reason or "Not available",
                 font=_SMALL, border=_BOX, align=_LEFT)
        # THE CARD'S OWN RENDERING beside the number, so a reader can check the
        # workbook against the screen without re-deriving anything, and so the
        # workbook and the PDF cannot disagree about precision.
        _put(ws, row, 3, entry.display or None, font=_VALUE, border=_BOX, align=_RIGHT)
        if entry.previous is not None:
            _put(ws, row, 4, entry.previous, font=_VALUE, border=_BOX, align=_RIGHT,
                 fmt=excel_number_format(entry.kind, currency))
        else:
            _put(ws, row, 4, None, border=_BOX)
        _put(ws, row, 5, entry.delta_display or None, font=_VALUE, border=_BOX, align=_RIGHT)
        _put(ws, row, 6, entry.trend or None, font=_VALUE, border=_BOX, align=_LEFT)
        # The wider-scope measurement the tile falls back to, when there is one.
        # Either the wider-scope fallback or this scope's own basis — see pdf.py.
        basis = (
            entry.measured_at
            if entry.measured_at
            else " · ".join(x for x in (entry.delta_basis, entry.evidence) if x)
        )
        _put(ws, row, 7, basis or None, font=_SMALL, border=_BOX, align=_LEFT)
        row += 1

    for i, width in enumerate((30, 18, 16, 18, 12, 10, 40), start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, width)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if section.note:
        row += 1
        _put(ws, row, 1, section.note, font=_SMALL, align=_WRAP)
        row += 1
    return row + 1


def _table_block(ws: Worksheet, row: int, table: Table, currency: str, *, freeze: bool) -> int:
    """One grid, with a frozen header row and an autofilter over the data."""
    if table.title:
        _put(ws, row, 1, table.title, font=_H2)
        row += 1

    for i, column in enumerate(table.columns, start=1):
        _put(ws, row, i, column.header, font=_TH, fill=_TH_FILL, border=_BOX,
             align=_RIGHT if is_numeric(column.kind) else _LEFT)
        ws.column_dimensions[get_column_letter(i)].width = column.width
    header_row = row
    row += 1

    for n, record in enumerate(table.rows):
        band = _BAND_FILL if n % 2 else None
        for i, column in enumerate(table.columns, start=1):
            value = record.get(column.key)
            numeric = is_numeric(column.kind) and isinstance(value, (int, float))
            _put(
                ws, row, i,
                value if (numeric or value is not None) else None,
                font=_VALUE,
                border=_BOX,
                fill=band,
                align=_RIGHT if numeric else _LEFT,
                fmt=excel_number_format(column.kind, currency) if numeric else None,
            )
        row += 1

    last = row - 1
    if table.rows:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(table.columns))}{last}"
        )
    if freeze:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if table.note:
        row += 1
        _put(ws, row, 1, table.note, font=_SMALL, align=_WRAP)
        row += 1
    return row + 1


def _kv_block(ws: Worksheet, row: int, section: Section) -> int:
    _put(ws, row, 1, section.title, font=_H2)
    row += 1
    for label, value in section.items:
        _put(ws, row, 1, str(label), font=_LABEL)
        _put(ws, row, 2, value, font=_VALUE, align=_WRAP)
        row += 1
    if section.note:
        _put(ws, row, 1, section.note, font=_SMALL, align=_WRAP)
        row += 1
    return row + 1


def _text_block(ws: Worksheet, row: int, section: Section) -> int:
    _put(ws, row, 1, section.title, font=_H2)
    row += 1
    for paragraph in section.items:
        _put(ws, row, 1, str(paragraph), font=_VALUE, align=_WRAP)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = max(15, 14 * (1 + len(str(paragraph)) // 110))
        row += 1
    return row + 1


def write(doc: ReportDoc, currency: str = "INR") -> bytes:
    """Render one report as `.xlsx` bytes.

    SHEET LAYOUT. The summary always leads. A section that names a `sheet` gets
    its own worksheet — that is how a detailed table stays usable instead of
    being buried under a summary — and everything else stacks onto the summary
    sheet in order.
    """
    book = Workbook()
    used: set[str] = set()

    summary = book.active
    summary.title = _sheet_name("Executive Summary", used)
    summary.sheet_view.showGridLines = False
    row = _cover(summary, doc, currency)

    for section in doc.sections:
        if section.sheet:
            sheet = book.create_sheet(_sheet_name(section.sheet, used))
            sheet.sheet_view.showGridLines = False
            at = 1
            if section.kind == "table" and section.table is not None:
                _put(sheet, at, 1, doc.brand, font=_BRANDLINE)
                _put(sheet, at + 1, 1, f"{doc.module} · {doc.scope_line}", font=_SMALL)
                at += 3
                _table_block(sheet, at, section.table, currency, freeze=True)
            elif section.kind == "kpi":
                _kpi_block(sheet, at, section, currency)
            elif section.kind == "text":
                _text_block(sheet, at, section)
            else:
                _kv_block(sheet, at, section)
            continue

        if section.kind == "kpi":
            row = _kpi_block(summary, row, section, currency)
        elif section.kind == "table" and section.table is not None:
            # THE SECTION'S OWN HEADING FIRST. `_table_block` writes the TABLE's
            # title, which is a subtitle describing the grid ("Trade spend by
            # promotion"); without this the section heading it belongs under
            # ("Promotion mix") never reached the sheet at all.
            if section.title and section.title != section.table.title:
                _put(summary, row, 1, section.title, font=_H2)
                row += 1
            row = _table_block(summary, row, section.table, currency, freeze=False)
        elif section.kind == "text":
            row = _text_block(summary, row, section)
        else:
            row = _kv_block(summary, row, section)

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
